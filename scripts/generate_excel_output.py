# 🕒 2025-12-17-19-36-00
# unified_data_dictionary/src/generate_excel_output.py
# Author: R. A. Carucci
# Purpose: Generate human-readable Excel documentation from unified data dictionary with multiple worksheets for fields, mappings, and validation rules

import json
import logging
from pathlib import Path
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelDictionaryGenerator:
    """Generate Excel documentation from unified dictionary."""
    
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        """Initialize Excel generator."""
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
            
    def load_dictionary(self, json_path: Path) -> Dict[str, Any]:
        """Load dictionary from JSON file."""
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
            
    def create_metadata_sheet(self, metadata: Dict):
        """Create metadata overview sheet."""
        ws = self.workbook.create_sheet("Metadata")
        
        # Title
        ws['A1'] = "Unified Data Dictionary - Metadata"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Metadata rows
        row = 3
        for key, value in metadata.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(value)
            row += 1
            
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
    def create_fields_sheet(self, fields: Dict):
        """Create main fields definition sheet."""
        ws = self.workbook.create_sheet("Fields")
        
        # Headers
        headers = [
            'Field Name',
            'Data Type',
            'Description',
            'Required',
            'Unique',
            'Example',
            'Source Systems',
            'Aliases',
            'PII Category',
            'Validation Rules'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
            
        # Data rows
        row = 2
        for field_name, field_def in sorted(fields.items()):
            ws.cell(row=row, column=1, value=field_name).border = self.BORDER
            ws.cell(row=row, column=2, value=field_def['data_type']).border = self.BORDER
            ws.cell(row=row, column=3, value=field_def['description']).border = self.BORDER
            ws.cell(row=row, column=4, value='Yes' if field_def['required'] else 'No').border = self.BORDER
            ws.cell(row=row, column=5, value='Yes' if field_def['unique'] else 'No').border = self.BORDER
            ws.cell(row=row, column=6, value=field_def.get('example', '')).border = self.BORDER
            ws.cell(row=row, column=7, value=', '.join(field_def['source_systems'])).border = self.BORDER
            ws.cell(row=row, column=8, value=', '.join(field_def.get('aliases', []))).border = self.BORDER
            ws.cell(row=row, column=9, value=field_def.get('pii_category', '')).border = self.BORDER
            ws.cell(row=row, column=10, value=', '.join(field_def.get('validation_rules', []))).border = self.BORDER
            
            # Wrap text for description
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
            
        # Auto-adjust column widths
        column_widths = [30, 12, 50, 10, 10, 20, 20, 25, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
            
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add auto-filter
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row-1}'
        
    def create_mappings_sheet(self, mappings: Dict):
        """Create field mappings sheet."""
        if not mappings:
            return
            
        ws = self.workbook.create_sheet("Mappings")
        
        # Headers
        headers = ['Mapping Name', 'Source Field', 'Target Field', 'Transformation']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
            
        # Data rows
        row = 2
        for mapping_name, mapping_data in sorted(mappings.items()):
            if isinstance(mapping_data, dict):
                for source_field, target_info in mapping_data.items():
                    ws.cell(row=row, column=1, value=mapping_name).border = self.BORDER
                    ws.cell(row=row, column=2, value=source_field).border = self.BORDER
                    
                    if isinstance(target_info, dict):
                        ws.cell(row=row, column=3, value=target_info.get('target_field', '')).border = self.BORDER
                        ws.cell(row=row, column=4, value=target_info.get('transformation', '')).border = self.BORDER
                    else:
                        ws.cell(row=row, column=3, value=str(target_info)).border = self.BORDER
                        
                    row += 1
                    
        # Auto-adjust column widths
        column_widths = [30, 30, 30, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
            
        # Freeze header row
        ws.freeze_panes = 'A2'
        
    def create_data_types_sheet(self, fields: Dict):
        """Create data types reference sheet."""
        ws = self.workbook.create_sheet("Data Types")
        
        # Count fields by data type
        type_counts = {}
        for field_def in fields.values():
            data_type = field_def['data_type']
            type_counts[data_type] = type_counts.get(data_type, 0) + 1
            
        # Headers
        headers = ['Data Type', 'Count', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
            
        # Data type descriptions
        type_descriptions = {
            'string': 'Text data of variable length',
            'integer': 'Whole numbers without decimals',
            'number': 'Numeric data including decimals',
            'boolean': 'True/False values',
            'datetime': 'Date and time combined',
            'date': 'Date only (YYYY-MM-DD)',
            'enum': 'Enumerated list of valid values'
        }
        
        # Data rows
        row = 2
        for data_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1, value=data_type).border = self.BORDER
            ws.cell(row=row, column=2, value=count).border = self.BORDER
            ws.cell(row=row, column=3, value=type_descriptions.get(data_type, '')).border = self.BORDER
            row += 1
            
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        
    def create_conflicts_sheet(self, conflicts: list):
        """Create conflicts resolution sheet."""
        if not conflicts:
            return
            
        ws = self.workbook.create_sheet("Conflicts")
        
        # Headers
        headers = ['Field Name', 'Conflict Type', 'Existing Value', 'New Value', 'Source System']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
            
        # Data rows
        row = 2
        for conflict in conflicts:
            field_name = conflict['field_name']
            source_system = conflict['source_system']
            
            for conflict_type, values in conflict['conflicts'].items():
                ws.cell(row=row, column=1, value=field_name).border = self.BORDER
                ws.cell(row=row, column=2, value=conflict_type).border = self.BORDER
                ws.cell(row=row, column=3, value=str(values['existing'])).border = self.BORDER
                ws.cell(row=row, column=4, value=str(values['new'])).border = self.BORDER
                ws.cell(row=row, column=5, value=source_system).border = self.BORDER
                row += 1
                
        # Column widths
        column_widths = [30, 15, 30, 30, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
            
    def generate(self, json_path: Path, output_path: Path):
        """Generate complete Excel documentation."""
        logger.info(f"Loading dictionary from {json_path}")
        dictionary = self.load_dictionary(json_path)
        
        # Create sheets
        logger.info("Creating Metadata sheet")
        self.create_metadata_sheet(dictionary['metadata'])
        
        logger.info("Creating Fields sheet")
        self.create_fields_sheet(dictionary['fields'])
        
        logger.info("Creating Mappings sheet")
        self.create_mappings_sheet(dictionary.get('mappings', {}))
        
        logger.info("Creating Data Types sheet")
        self.create_data_types_sheet(dictionary['fields'])
        
        logger.info("Creating Conflicts sheet")
        self.create_conflicts_sheet(dictionary.get('conflicts', []))
        
        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
        logger.info(f"Excel documentation saved to {output_path}")


def main():
    """Main Excel generation workflow."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate Excel dictionary documentation")
    parser.add_argument(
        '--input',
        type=Path,
        default=Path('output/unified_data_dictionary.json'),
        help='Input JSON dictionary file'
    )
    parser.add_argument(
        '--output',
        type=Path,
        default=Path('output/unified_data_dictionary.xlsx'),
        help='Output Excel file'
    )
    
    args = parser.parse_args()
    
    generator = ExcelDictionaryGenerator()
    generator.generate(args.input, args.output)
    
    logger.info("Excel generation complete")


if __name__ == '__main__':
    main()
